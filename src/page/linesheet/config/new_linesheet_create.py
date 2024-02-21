#--- Generate
# #### Function Hear
# import mysql.connector
from sqlalchemy import create_engine
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule, CellIsRule
import openpyxl
from openpyxl.styles import PatternFill





def convert_hex_to_argb(hex_code):
    if hex_code!="" and hex_code!=None and hex_code!="None" and len(hex_code)==7:
        # Remove the '#' character if present
        hex_code = hex_code.lstrip('#')

        alpha_hex = 'FF'

        argb_color = f'{alpha_hex}{hex_code}'
        return argb_color
    else:
        alpha_hex = 'FF'
        return f'{alpha_hex}FFFFFF'
# ===========
def get_string_location(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return '$'+openpyxl.utils.get_column_letter(column) + '$'+str(row)
# Add the condition formatting
# ============
def apply_conditional_formatting_by_header(workbook, worksheet, header_value, condition, fill_color):
    # Find the column number where the header_value is located
    column_number = None
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == header_value:
                column_number = cell.column_letter
                break
        if column_number:
            break

    if column_number is None:
        # print(f"Header '{header_value}' not found in the worksheet.")
        return

    # Create a conditional formatting rule
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    rule = CellIsRule(operator="equal",formula=['="'+condition+'"'], fill=fill)

    # Add the rule to the worksheet
    worksheet.conditional_formatting.add(f'{column_number}1:{column_number}{worksheet.max_row}', rule)

# ===========
# Add the columns to the worksheet, starting at cell A1
# ===========

def add_column(row,color,background,lists,wrapText,border,comment,author,allow_comment):
    for i, value in enumerate(lists):
        # Identity cell
        cell = worksheet.cell(row=row, column=i+1)
        cell.value = value
        # Setup cell font color
        cell.font = openpyxl.styles.Font(color=color)
        # Setup cell background color
        cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=background)
        # Setup Alignment at center for horizontal and vertical
        cell.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=wrapText)
        # Add a comment
        if comment[i] is not None and allow_comment==True and comment[i] !="":
            cell.comment = openpyxl.comments.Comment(text=comment[i], author='script')
        # Setup border
        if border == True or border == "True":
            cell.border = openpyxl.styles.Border(
                top=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                right=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                bottom=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                left=openpyxl.styles.Side(color='F0F0F0', style='thin')
            )
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].auto_size
# Add the columns to the worksheet, starting at cell A1
# ===========
def add_header(row,color,background,lists,bold,prefix = '',merge_cell = False,border=True):
    wrapText = False
    for i, value in enumerate(lists):
        # Identity cell
        cell = worksheet.cell(row=row, column=i+1)
        # Setup cell value
        if lists[i] != lists[i-1] :
            cell.value = str(prefix) + str(value)
            #merge
            if merge_group[i] == True and merge_cell == True:
                # check cell in same group
                num_cell = lists.count(value)
                # merge
                target_string = get_string_location(row-1, i) + ':'+ get_string_location(row, i+num_cell-1)
                workbook['IM_FORM'].merge_cells(target_string)
                wrapText = True
            else:
                wrapText = False
             # Setup border
            if border == True:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(color='F0F0F0', style='thin')
                )
                worksheet.cell(row=row+1, column=i+1).border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(color='F0F0F0', style='thin')
                )
        # Setup cell font color
        cell.font = openpyxl.styles.Font(color=color,bold=bold)
        # Setup cell background color
        cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=background)
        # Setup Alignment at center for horizontal and vertical
        cell.alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center',wrapText=wrapText,indent=1)
# ===========
def add_background_color(workbook, sheet_name, start_column, end_column, row, color):
    # Select the sheet where you want to add the background color
    sheet = workbook[sheet_name]
    # Set the color of the cells
    for col in range(start_column, end_column + 1):
        cell = sheet.cell(column=col, row=row)
        cell.fill = openpyxl.styles.PatternFill(fgColor=color, patternType='solid')
# ===========
def create_define_name(workbook ,sheet_name, attribute_code,length_list,column_num):
    string_range = get_string_location(1,column_num)+':'+get_string_location(length_list,column_num)
    new_range = openpyxl.workbook.defined_name.DefinedName(attribute_code, attr_text=sheet_name+'!'+string_range)

    try:
        workbook.defined_names.append(new_range)
    except:
        workbook.defined_names.add(new_range)


# ===========
def add_list_of_attribute(workbook, sheet_name,option_lists,attribute_code):
    sheet = workbook[sheet_name]
    col = sheet.max_column + 1
    sheet.cell(column=col, row=1, value=attribute_code)
    for i ,value in enumerate(option_lists):
        sheet.cell(column=col, row=i+2, value=value)
    #create definition name
    create_define_name(workbook ,sheet_name, attribute_code,len(option_lists),col-1)
# ===========
def add_dropdown(workbook, sheet_name, column, start_row, end_row, options):
    # Select the sheet where you want to add the dropdown
    sheet = workbook[sheet_name]
    # Create a DataValidation object with the dropdown options

    if options =="sub_dept":
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+"=OFFSET(DEPT_SUBDEPT_MAPPING!$D$1,MATCH(@INDIRECT(ADDRESS(ROW(), COLUMN()-2)),DEPT_SUBDEPT_MAPPING!$D:$D,0)-1,2,COUNTIF(DEPT_SUBDEPT_MAPPING!$D:$D,@INDIRECT(ADDRESS(ROW(), COLUMN()-2))),1)", allow_blank=True)
    elif options =="class":
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+"=OFFSET(DEPT_SUBDEPT_MAPPING!$F$1,MATCH(@INDIRECT(ADDRESS(ROW(), COLUMN()-2)),DEPT_SUBDEPT_MAPPING!$F:$F,0)-1,2,COUNTIF(DEPT_SUBDEPT_MAPPING!$F:$F,@INDIRECT(ADDRESS(ROW(), COLUMN()-2))),1)", allow_blank=True)
    elif options =="sub_class":
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+"=OFFSET(DEPT_SUBDEPT_MAPPING!$H$1,MATCH(@INDIRECT(ADDRESS(ROW(), COLUMN()-2)),DEPT_SUBDEPT_MAPPING!$H:$H,0)-1,2,COUNTIF(DEPT_SUBDEPT_MAPPING!$H:$H,@INDIRECT(ADDRESS(ROW(), COLUMN()-2))),1)", allow_blank=True)
    elif options =="size":
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+"=OFFSET(JDA_SIZE_MAPPING!$A$1,MATCH(@INDIRECT(ADDRESS(ROW(), COLUMN()-2)),JDA_SIZE_MAPPING!$A:$A,0)-1,1,COUNTIF(JDA_SIZE_MAPPING!$A:$A,@INDIRECT(ADDRESS(ROW(), COLUMN()-2))),1)", allow_blank=True)
    elif options =="color_group":
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+"=OFFSET(COLOR_MAPPING!$C$1,1,0,COUNTA(COLOR_MAPPING!$C:$C)-1,1)", allow_blank=True)
    elif options =="color_shade":
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+"=OFFSET(COLOR_MAPPING!$C$1,MATCH(@INDIRECT(ADDRESS(ROW(), COLUMN()-1)),COLOR_MAPPING!$C:$C,0)-1,1,COUNTIF(COLOR_MAPPING!$C:$C,@INDIRECT(ADDRESS(ROW(), COLUMN()-1))),1)", allow_blank=True)
    elif options =="shoe_size":
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+"=OFFSET(COMMON_SIZE_MAPPING!$A$1,MATCH(@INDIRECT(ADDRESS(1, COLUMN()))&@INDIRECT(ADDRESS(ROW(), COLUMN()-1)),COMMON_SIZE_MAPPING!$A:$A,0)-1,3,COUNTIF(COMMON_SIZE_MAPPING!$A:$A,@INDIRECT(ADDRESS(1, COLUMN()))&@INDIRECT(ADDRESS(ROW(), COLUMN()-1))),1)", allow_blank=True)
    elif options =="clothing_size":
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+"=OFFSET(COMMON_SIZE_MAPPING!$A$1,MATCH(@INDIRECT(ADDRESS(1, COLUMN()))&@INDIRECT(ADDRESS(ROW(), COLUMN()-1)),COMMON_SIZE_MAPPING!$A:$A,0)-1,3,COUNTIF(COMMON_SIZE_MAPPING!$A:$A,@INDIRECT(ADDRESS(1, COLUMN()))&@INDIRECT(ADDRESS(ROW(), COLUMN()-1))),1)", allow_blank=True)
    else:
        validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+options, allow_blank=True)
    # Add the DataValidation object to the sheet
    sheet.add_data_validation(validation)
    # Add the DataValidation object to the specified range of cells
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(column=column, row=row)
        validation.add(cell)
# ===========
def add_formula(workbook, sheet_name, column, start_row, end_row, formula):
    # Select the sheet where you want to add the dropdown
    sheet = workbook[sheet_name]
    # Add the DataValidation object to the specified range of cells
    for row in range(start_row, end_row + 1):
        sheet.cell(column=column, row=row, value="="+str(formula))
        # add background
        cell = sheet.cell(column=column, row=row)
        cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='E7E7E7')
        cell.border = openpyxl.styles.Border(
                        top=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                        right=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                        bottom=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                        left=openpyxl.styles.Side(color='F0F0F0', style='thin')
                    )
        # Setup Alignment at center for horizontal and vertical
        cell.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=False)
import re
import string
import datetime
import random
# ===========
def generate_track_id():
    current_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    random_number = random.randint(0, 999)
    track_id = current_time + str(random_number).zfill(3)
    return track_id
# ===========
def clean_string(string):
    """Replace special characters and convert a string to lowercase."""
    string = string.lower()
    return string
# ===========
def convert_sale_online_query(sale_channel):
    if sale_channel=='offline':
            return  "and sale_channel like '%offline%'"
    elif sale_channel=='online':
            return  "and sale_channel like '%online%'"
    else :
            return  ""
def convert_gsheets_url(u):
    try:
        worksheet_id = u.split('#gid=')[1]
    except:
        # Couldn't get worksheet id. Ignore it
        worksheet_id = None
    u = re.findall('https://docs.google.com/spreadsheets/d/.*?/',u)[0]
    u += 'export'
    u += '?format=csv'
    if worksheet_id:
        u += '&gid={}'.format(worksheet_id)
    return u
def row_contains_excluded(row):
    return not any(row[col] in values_to_exclude for col in column_names)
#--- Generate functions
# ===========
def generate_form(brand,template,sku,launch_date,stock_source,sale_channel,production_type,contact_person):

    track_id = generate_track_id()
    # ### Read data ####################################################
    # this session use from get a configurable data from the database , so this will include
    # 1. attribute_setting
    # 2. attribute option setting
    # replace to multi template  #####################################

    filter_template_cat_sql = template.replace(",", "','")
    filter_template_header_sql = template.replace(",", " not in ('N','AR','AO','unused') or ")
    selected_template_list = list(map(str,template.split(",")))
    selected_sale_channel_list = list(map(str,sale_channel.split(",")))

    # - Connect to the database
    #get source of mapping
    # #UAT
    # index_source = "https://docs.google.com/spreadsheets/d/18bS_SQWfb0tcuP0LywyfIot1bE_Rt7dV9qoSfzGtBsw/edit#gid=1054033513"
    #PROD
    index_source = "https://docs.google.com/spreadsheets/d/11cAVRnwQbD2LeCjL3nWuHozYzLF27CcwgyJnk82ixWc/edit#gid=1370721427"
    url = convert_gsheets_url(index_source)
    index = pd.read_csv(url)

    attribute_setting_url = index[index['sheet_name'] == 'attribute_setting']['url'].values[0]
    attribute_option_url = index[index['sheet_name'] == 'attribute_option']['url'].values[0]
    categories_mapping_url = index[index['sheet_name'] == 'categories_mapping']['url'].values[0]
    shipping_mapping_url = index[index['sheet_name'] == 'shipping_mapping']['url'].values[0]
    color_mapping_url = index[index['sheet_name'] == 'color_mapping']['url'].values[0]
    dept_subdept_mappping_url = index[index['sheet_name'] == 'dept_subdept_mappping']['url'].values[0]
    jda_size_mapping_url = index[index['sheet_name'] == 'jda_size_mapping']['url'].values[0]
    datapump_store_mapping_url = index[index['sheet_name'] == 'datapump_store_mapping']['url'].values[0]
    common_size_mapping_url = index[index['sheet_name'] == 'common_size_mapping']['url'].values[0]
    brand_store_mapping_url = index[index['sheet_name'] == 'brand_store_mapping']['url'].values[0]

     # - get common size mapping
    brand_store_mapping_query = convert_gsheets_url(brand_store_mapping_url)
    brand_store_mapping = pd.read_csv(brand_store_mapping_query)

    # - get common size mapping
    common_size_mapping_query = convert_gsheets_url(common_size_mapping_url)
    common_size_mapping = pd.read_csv(common_size_mapping_query)

    # - get header linesheet

    attribute_setting = convert_gsheets_url(attribute_setting_url)
    attribute = pd.read_csv(attribute_setting)
    attribute = attribute[attribute['status']=='Actived']
    attribute_original=attribute

    attribute['sale_channel'] = attribute['sale_channel'].str.split(',')
    attribute = attribute[attribute['sale_channel'].apply(lambda x: any(channel in selected_sale_channel_list for channel in x))]


    values_to_exclude=["N","AR","AO"]
    attribute_query_string = ' or '.join([f'{col} not in {values_to_exclude}' for col in selected_template_list])
    attribute = attribute.query(attribute_query_string)
    attribute = attribute.drop_duplicates(subset=['linesheet_code'])
    attribute = attribute.fillna("")


    # - get  store
    datapump_store_query = convert_gsheets_url(datapump_store_mapping_url)
    datapump_store = pd.read_csv(datapump_store_query)
    datapump_store['linesheet_code']='store_stock'
    datapump_store['input_option']=datapump_store['system_label']
    datapump_store = datapump_store[['linesheet_code','input_option']]

    # - get dropdown
    attribute_options_query = convert_gsheets_url(attribute_option_url)
    attribute_options = pd.read_csv(attribute_options_query, dtype='str')
    attribute_options = attribute_options.fillna("")
    attribute_options["code_lookup"] = attribute_options['linesheet_code']+attribute_options['input_option']
    attribute_options["label_lookup"] = attribute_options['linesheet_code']+attribute_options['option_en']
    lookup_sheet = attribute_options[['linesheet_code','code_lookup','label_lookup','input_option','option_en']]
    attribute_options = attribute_options[['linesheet_code','input_option']]
    attribute_options = attribute_options.drop_duplicates()
    attribute_options = attribute_options.fillna("")




    # - get categories

    # query = "SELECT label_th FROM im_form.categories_setting where deepen_level = 1 and family in ('"+filter_template_cat_sql +"')"
    # categories_setting = pd.read_sql_query(query, engine)
    categories_setting_query = convert_gsheets_url(categories_mapping_url)
    categories_setting = pd.read_csv(categories_setting_query, dtype='str')
    categories_setting=categories_setting[categories_setting['family'].isin(selected_template_list)]
    categories_setting=categories_setting[categories_setting['deepen_level']=="TRUE"]
    categories_setting=categories_setting[['label_th']]
    categories_setting = categories_setting.fillna("")

    # - get categories family template


    family_setting_query = convert_gsheets_url(attribute_setting_url)
    family_setting = pd.read_csv(family_setting_query , dtype='str')
    family_setting = family_setting.fillna("")
    # family_setting = attribute_original
    columns_to_drop = ['id',
                        'information_type',
                        'status',
                        'enhancement',
                        'specific_brand',
                        'field_label',
                        'field_type',
                        'both_language',
                        'scopable',
                        'description',
                        'tool_tips',
                        'session',
                        'sub_session',
                        'merge_group',
                        'sale_channel',
                        'formula',
                        'pim_code',
                        'pim_code_hard_header',
                        'convertor_function',
                        'linesheet_code_unit',
                        'label_desc_en',
                        'label_desc_th',
                        'value_desc_format',
                        'sort_bullet_point',
                        'grouping_common',
                        'owner_field']

    family_setting = family_setting.drop(columns_to_drop, axis=1)
    # family_setting = family_setting[['linesheet_code', 'auto__motorcycle_supplies', 'baby_feeding', 'bath_body', 'bedding', 'books', 'camping_equipments', 'clothing', 'computers', 'console_gaming', 'cooking_dining', 'desk_phone', 'fans_air_purifiers', 'fashion_accessory', 'fragrance', 'furniture', 'gadgets', 'gaming', 'gift_card', 'groceries', 'hair_care', 'health_care', 'home_decoration', 'home_equipment_supplies', 'hobby', 'luggages', 'large_appliances', 'makeup', 'makeup_tools', 'mobile_tablets', 'nails', 'nails_tools', 'personal_care', 'pet_equipment_supplies', 'skincare', 'small_appliances', 'sports_accessory', 'sports_equipments', 'stationery', 'swimwear', 'television', 'toolings', 'toys', 'travel_accessories', 'underwear', 'watches']]

    family_setting = family_setting.fillna("")
    family_setting = family_setting.astype(str)


    categories_family_setting_query = convert_gsheets_url(categories_mapping_url)
    categories_family_setting = pd.read_csv(categories_family_setting_query, dtype='str')
    categories_family_setting = categories_family_setting[categories_family_setting['deepen_level']=="TRUE"]
    categories_family_setting = categories_family_setting[['label_th','family']]
    categories_family_setting = categories_family_setting.fillna("")


    # - get dept_subdept
    # query = "SELECT * FROM im_form.dept_subdept_mapping;"
    # dept_subdept_mapping = pd.read_sql_query(query, engine)
    dept_subdept_mapping_query = convert_gsheets_url(dept_subdept_mappping_url)
    dept_subdept_mapping = pd.read_csv(dept_subdept_mapping_query, dtype='str')
    dept_subdept_mapping=dept_subdept_mapping.fillna("None")

    jda_size_mapping_query = convert_gsheets_url(jda_size_mapping_url)
    jda_size_mapping = pd.read_csv(jda_size_mapping_query, dtype='str')
    jda_size_mapping = jda_size_mapping.fillna("None")


    color_mapping_query = convert_gsheets_url(color_mapping_url)
    color_mapping = pd.read_csv(color_mapping_query, dtype='str')
    color_mapping = color_mapping.fillna("None")

    # - Merge categories to options
    categories_setting['linesheet_code'] = 'online_categories'
    # - Rename the column headers
    categories_setting = categories_setting.rename(columns={'label_th': 'input_option'})
    attribute_options = pd.concat([attribute_options, categories_setting], axis=0)

    attribute_options = pd.concat([attribute_options, datapump_store], axis=0)





    # - Close the connection
    # cnx.close()
# #### @ Create an config backup file ##########################3
    import os

    os.makedirs("linesheet" , exist_ok=True)

    global codes
    global labels
    global indicators
    global comment
    global group_header
    global group_sub_header
    global merge_group
    global translation
    global fill_type
# create column indicator
    attribute["column_indication"] = attribute.loc[:, selected_template_list].apply(lambda x: ' '.join(x.astype(str)), axis=1)
# convert Data Frame to a list
    codes = attribute['linesheet_code'].values.tolist()
    labels = attribute['field_label'].values.tolist()
    indicators = attribute['column_indication'].values.tolist()
    owner_field = attribute['owner_field'].values.tolist()
    comment = attribute['tool_tips'].values.tolist()
    group_header = attribute['session'].values.tolist()
    group_sub_header = attribute['sub_session'].values.tolist()
    merge_group =  attribute['merge_group'].values.tolist()
    translation = attribute['both_language'].values.tolist()
    formula = attribute['formula'].values.tolist()
    fill_type = attribute['field_type'].values.tolist()
# ##### Add translations column
    i = 0
    for j, value in enumerate(translation):
        if value == 1 or value == True or value == "True":
            codes.insert(i,codes[i]+'_en')
            labels.insert(i,labels[i]+' [English]')
            codes[i+1] = codes[i+1]+'_th'
            labels[i+1] = labels[i+1]+' [Thai]'
            indicators.insert(i,indicators[i])
            comment.insert(i,comment[i])
            owner_field.insert(i,owner_field[i])
            group_header.insert(i,group_header[i])
            group_sub_header.insert(i,group_sub_header[i])
            merge_group.insert(i,merge_group[i])
            formula.insert(i,formula[i])
            fill_type.insert(i,fill_type[i])
            i += 1
        i += 1
# ### Create a sheet
    # This function use for create a excel file , so this one include the sheet below
    global workbook
    global worksheet
    global worksheet_form
    global ws_att_option
    global ws_inlink_data
    global ws_color_mapping_sheet
    # - Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    workbook.properties.author = "Copyright © Central Online"
    # Change the name of the worksheet
    worksheet.title = 'IM_FORM'
    worksheet_form = workbook['IM_FORM']
    # Create a new sheet with the name "Sheet2"
    ws_att_option = workbook.create_sheet("ATT_OPTION")
    ws_inlink_data = workbook.create_sheet("IN_LINK_DATA")
    ws_lookup_sheet = workbook.create_sheet("LOOKUP_SHEET")
    ws_depsubdept_mapping_sheet = workbook.create_sheet("DEPT_SUBDEPT_MAPPING")
    ws_family_template = workbook.create_sheet("FAMILY_TEMPLATE")
    ws_jda_size_mapping_sheet=workbook.create_sheet("JDA_SIZE_MAPPING")
    ws_color_mapping_sheet=workbook.create_sheet("COLOR_MAPPING")
    ws_datapump_store_sheet=workbook.create_sheet("DATAPUMP_STORE_MAPPING")
    ws_common_size_mapping_sheet=workbook.create_sheet("COMMON_SIZE_MAPPING")
    ws_brand_store_mapping_sheet=workbook.create_sheet("BRAND_STORE_MAPPING")

    # Hide the sheet
    ws_att_option.sheet_state = 'hidden'
    ws_inlink_data.sheet_state = 'hidden'
    ws_lookup_sheet.sheet_state = 'hidden'
    ws_depsubdept_mapping_sheet.sheet_state = 'hidden'
    ws_family_template.sheet_state = 'hidden'
    ws_jda_size_mapping_sheet.sheet_state = 'hidden'
    ws_color_mapping_sheet.sheet_state = 'hidden'
    ws_datapump_store_sheet.sheet_state = 'hidden'
    ws_common_size_mapping_sheet.sheet_state = 'hidden'
    ws_brand_store_mapping_sheet.sheet_state = 'hidden'

# #### General information's
# - Add version information
    worksheet.cell(row=2, column=2).value = 'New Omni linesheet Version ONE'
    # worksheet.cell(row=3, column=2).value = 'Template : ' +str(template)
    worksheet.cell(row=2, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=True)
    worksheet.cell(row=3, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)

# Add lookup sheet
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(lookup_sheet, index=True, header=True):
        ws_lookup_sheet.append(r)

# Add dept subdept mapping
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(dept_subdept_mapping, index=False, header=True):
        ws_depsubdept_mapping_sheet.append(r)

# Add jda size mapping
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(jda_size_mapping, index=False, header=True):
        ws_jda_size_mapping_sheet.append(r)

# Add color mapping
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(color_mapping, index=False, header=True):
        ws_color_mapping_sheet.append(r)

# Add datapump store mapping
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(datapump_store, index=False, header=True):
        ws_datapump_store_sheet.append(r)

# Add common_size mapping
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(common_size_mapping, index=False, header=True):
        ws_common_size_mapping_sheet.append(r)

# Add brand_store_mapping
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(brand_store_mapping, index=False, header=True):
        ws_brand_store_mapping_sheet.append(r)


# Add family template sheet

    # Get unique column names (excluding 'linesheet_code')
    try:
        column_names = family_setting.columns[1:]
    except KeyError as err:
        print(err)
    # column_names = family_setting.columns
    # Create a dictionary to store the templates
    templates = {}
    # Iterate over each column name
    for col in column_names:
        template_values = {'R': '', 'O': '', 'N': '', 'AO': '', 'AR': '', "": ''}
        for index, value in enumerate(family_setting[col]):
            template_values[value] += family_setting['linesheet_code'][index] + ','
        templates[col] = template_values
    # Create a new DataFrame for the templates
    templates_df = pd.DataFrame(templates)
    # Reset index and rename index column
    templates_df = templates_df.reset_index().rename(columns={'index': 'family'})
    # Transpose the DataFrame
    templates_df = templates_df.transpose()
    templates_df.columns = templates_df.iloc[0]
    templates_df = templates_df[1:]
    # Reset index and rename index column
    templates_df = templates_df.reset_index().rename(columns={'index': 'family'})
    categories_family_mapping = pd.merge(categories_family_setting, templates_df, on='family', how='left')
    categories_family_mapping.rename(columns={'label_th': 'categories'}, inplace=True)

    for r in dataframe_to_rows(categories_family_mapping, index=False, header=True):
        ws_family_template.append(r)

# Add inLink Data
    # -- get current date time
    from datetime import datetime
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    ws_inlink_data.append(['attribute', 'value'])
    ws_inlink_data.append(['brand', brand])
    ws_inlink_data.append(['sku', sku])
    ws_inlink_data.append(['stock_source', "=IM_FORM!C6"])
    ws_inlink_data.append(['template', template])
    ws_inlink_data.append(['sale_channel', sale_channel])
    ws_inlink_data.append(['launch_date', launch_date])
    ws_inlink_data.append(['production_type', production_type])
    ws_inlink_data.append(['form_version', '2.3.2'])
    ws_inlink_data.append(['create_date', dt_string])
    ws_inlink_data.append(['author', 'system'])
    ws_inlink_data.append(['track_id',track_id])
    ws_inlink_data.append(['check_require', 'unavailable'])
    ws_inlink_data.append(['check_list_validations', 'unavailable'])
    ws_inlink_data.append(['contact_person', contact_person])
    ws_inlink_data.append(['msg_validate_mandatory_checking', "Not validate yet"])
    ws_inlink_data.append(['msg_validate_type_checking', "Not validate yet"])
# ##### identity row
    row_field_type=9
    row_session=10
    row_sub_session=11
    row_owner_field=12
    row_mandatory_badge=13
    row_header=14
    row_begin_form=15


# ##### Add a Table information's
    # Set the table range
    range_string = 'B3:C7'
    # Create a new table
    table = openpyxl.worksheet.table.Table(displayName="general_information", ref=range_string)
    # Add the table to the worksheet
    worksheet_form.add_table(table)
    # add table information
    ## add column
    worksheet_form.cell(row=3, column=2).value = 'General information'
    worksheet_form.cell(row=3, column=3).value = 'Value'
    ## add Row launch date
    worksheet_form.cell(row=4, column=2).value = 'Launch date'
    worksheet_form.cell(row=4, column=3).value = '=IN_LINK_DATA!B7' #launch_date
    ## add Row Production type
    worksheet_form.cell(row=5, column=2).value = 'Production type'
    worksheet_form.cell(row=5, column=3).value = '=IN_LINK_DATA!B8' #production_type
    ## add stock_source stock
    worksheet_form.cell(row=6, column=2).value = 'Multi Stock source'
    worksheet_form.cell(row=6, column=2).comment = openpyxl.comments.Comment(text="Multi Select สามารถเลือกได้มากว่า 1 store", author='script')
    # worksheet_form.cell(row=6, column=3).value = '=IN_LINK_DATA!B4' #stock_source

    worksheet.cell(row=7, column=2).value = 'Template'
    worksheet.cell(row=7, column=3).value = '=IN_LINK_DATA!B5'

    worksheet_form.cell(row=3, column=5).value = 'Contact person'
    worksheet_form.cell(row=4, column=5).value = '=IN_LINK_DATA!B15'

    worksheet_form.cell(row=3, column=2).font = openpyxl.styles.Font(color='C00000',bold=True)

    worksheet_form.cell(row=3, column=3).font = openpyxl.styles.Font(color='C00000',bold=False)

    worksheet_form.cell(row=3, column=2).fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='BFBFBF')
    worksheet_form.cell(row=3, column=3).fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='BFBFBF')

    worksheet_form.cell(row=4, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)
    worksheet_form.cell(row=4, column=3).font = openpyxl.styles.Font(color='FFFF25',bold=False)
    worksheet_form.cell(row=5, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)
    worksheet_form.cell(row=5, column=3).font = openpyxl.styles.Font(color='FFFF25',bold=False)
    worksheet_form.cell(row=6, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)
    worksheet_form.cell(row=6, column=3).font = openpyxl.styles.Font(color='FFFF25',bold=False)
    worksheet_form.cell(row=7, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)
    worksheet_form.cell(row=7, column=3).font = openpyxl.styles.Font(color='FFFF25',bold=False)

    worksheet_form.cell(row=3, column=5).font = openpyxl.styles.Font(color='C00000',bold=False)
    worksheet_form.cell(row=4, column=5).font = openpyxl.styles.Font(color='FFFF25',bold=False)

    worksheet_form.cell(row=3, column=5).fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='BFBFBF')
    worksheet_form.cell(row=3, column=6).fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='BFBFBF')


    #  #create_button for validate macro
    # worksheet_form.cell(row=4, column=6).value = "Validate Information"
    # worksheet_form.cell(row=4, column=6).font = openpyxl.styles.Font(color='FFFFFF')
    # worksheet_form.cell(row=4, column=6).font = openpyxl.styles.Font(bold=False)
    # worksheet_form.cell(row=4, column=6).fill = openpyxl.styles.PatternFill(bgColor="3A3838")
    # worksheet_form.cell(row=4, column=7).fill = openpyxl.styles.PatternFill(bgColor="3A3838")
    # worksheet_form.cell(row=4, column=8).fill = openpyxl.styles.PatternFill(bgColor="3A3838")
    # worksheet_form.cell(row=4, column=9).fill = openpyxl.styles.PatternFill(bgColor="3A3838")

# #### Convert require field to indicator
    # change indicator
    for i in range(len(indicators)):
        if "R" in indicators[i]:
            indicators[i] ="Require ✔"
        else:
            indicators[i] = ""
        # indicators[i] = indicators[i].replace("R", "✔")
        # indicators[i] = indicators[i].replace("R", "Require")
        # indicators[i] = indicators[i].replace("O", "")
    # change boolean
    for i in range(len(merge_group)):
        if merge_group[i] == 1 or merge_group[i]  == True or merge_group[i] =="True":
            merge_group[i] = True
        else:
            merge_group[i] = False
# #### call function Add a column
    # call add column function



    add_column(1,'C91818','B41010',codes,False,False,comment,'script',False)
    add_column(row_header,'434343','E5E5E5',labels,True,True,comment,'script',True)
    add_column(row_mandatory_badge,'F73434','BFCAD0',indicators,False,True,comment,'script',False)
    add_column(row_field_type,'D9D9D9','B41010',fill_type,False,False,comment,'script',False)
    add_column(row_owner_field,'434343','BFCAD0',owner_field,False,True,comment,'script',False)

    # Get the length of the list
    latest_column = len(codes)

    # custom bg color for fill type
    for column in range(1,latest_column+1):
        field_type = worksheet_form.cell(row=row_field_type, column=column).value
        if field_type=='Formula':
            bg_color='808080'
        elif field_type=='Number only':
            bg_color='31869B'
        elif field_type=='link':
            bg_color='43A8EC'
        elif field_type=='AUTOMATION':
            bg_color='808080'
        elif field_type=='Free Text':
            bg_color='14C671'
        elif field_type=='Simple Select':
            bg_color='A7A121'
        elif field_type=='Boolean':
            bg_color='A7A121'
        elif field_type=='Multiple Select':
            bg_color='CD7515'
        elif field_type=='Automation':
            bg_color='808080'


        add_background_color(workbook, 'IM_FORM', column, column, row_field_type, bg_color)

    # #### Add header
    add_header(row_session,'F0F0F0','161616',group_header,True,merge_cell=True,border=True)
    add_header(row_sub_session,'F0F0F0','161616',group_sub_header,False,'▶ ',merge_cell=True,border=False)
    # #### Add row background space
    start_row = 1
    end_row = 8
    # call function
    for row in range(start_row,end_row+1):
        add_background_color(workbook, 'IM_FORM', 1, latest_column, row, 'B41010')
    # #### Set the width of columns
    # Set the width of all columns to fit the contents of the cells
    for column_index in range(1,  latest_column+ 1):
        column = openpyxl.utils.get_column_letter(column_index)
        worksheet.column_dimensions[column].auto_size = True
## Add a dropdown list
    # - Loop add a dropdown
    # - Loop check value in each column and create a list of option
    start_col = 1
    for col in range(start_col, latest_column + 1):
        column_code = worksheet.cell(column=col, row=1).value
        option_frame = attribute_options[attribute_options['linesheet_code']==column_code]
        option_lists = option_frame['input_option'].values.tolist()
        if len(option_lists) != 0 :
            # Update the list of attribute in the sheet
            add_list_of_attribute(workbook, "ATT_OPTION", option_lists,column_code)
            # add data validation and create a dropdown
            add_dropdown(workbook,'IM_FORM', col, row_begin_form, int(sku)+row_begin_form, column_code)
            option_lists.clear()
    #add formula
    for col in range(start_col, latest_column + 1):
        if formula[col-1]!=""  and formula[col-1]!=None:
           add_formula(workbook,'IM_FORM', col, row_begin_form, int(sku)+row_begin_form, formula[col-1])

##a dd dropdown for store stock
    option_frame = attribute_options[attribute_options['linesheet_code']=='store_stock']
    option_lists = option_frame['input_option'].values.tolist()
    add_list_of_attribute(workbook, "ATT_OPTION", option_lists,'store_stock')
    add_dropdown(workbook,'IM_FORM', 3, 6, 6, 'store_stock')
    option_lists.clear()


####### convert to number

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for column in sheet.iter_cols(min_row=2):
            for cell in column:
                try:
                    cell.value = float(cell.value)
                except (ValueError, TypeError):
                    pass

####### create condition formatting for color

    for row in ws_color_mapping_sheet.iter_rows(min_row=2):
        color_name = row[3].value  # Access the cell in the fourth column (column 4) for color_name
        hex_code = row[7].value     # Access the cell in the eighth column (column 8) for hex_code


        apply_conditional_formatting_by_header(workbook, worksheet_form, 'color_shade', color_name, convert_hex_to_argb(hex_code))
        apply_conditional_formatting_by_header(workbook, ws_color_mapping_sheet, 'input_option', color_name, convert_hex_to_argb(hex_code))

# ##### Save a file
    file_name = 'linesheet/'+'CDS_LINESHEET_'+str(brand).upper()+'_'+str(sku)+'_SKUs_TID_'+track_id+'.xlsm'
    workbook.save(file_name)
    from openpyxl import Workbook
    from openpyxl import load_workbook
    wb1 = load_workbook(file_name)
    wb2 = load_workbook(file_name, keep_vba=True)
    wb2.save(file_name)
# Add vba multi-select
    import os
    current_directory = os.getcwd()
    # current_directory ='linesheet'
    files = os.listdir(current_directory)
    import win32com.client
    # Open the Excel file
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AutomationSecurity =1
    # xl.AutomationSecurity = 3

    wb = xl.Workbooks.Open(current_directory+'\\linesheet\\CDS_LINESHEET_'+str(brand).upper()+'_'+str(sku)+'_SKUs_TID_'+track_id+'.xlsm')
    # Access the sheet
    sheet = wb.Sheets("IM_FORM")
    # Access the VBA project
    vb = wb.VBProject
    # Create a new module
    module = vb.VBComponents.Add(1)
    module.Name = "im_form"
    code_name = sheet.codeName
    excelModule = wb.VBProject.VBComponents(code_name)
    # Insert the macro code


    import os

    with open(os.path.join(os.path.dirname(__file__), 'multi_select_cell.vba'), 'r') as f:
        code = f.read()
    # module.CodeModule.AddFromString(code)
    excelModule.CodeModule.AddFromString(code)
    # Add the macro to the sheet
    # sheet.CodeName = "Sheet1"
    sheet.OnSheetActivate = "im_form.Worksheet_Change"







    filename = 'linesheet/'+'CDS_LINESHEET_'+str(brand).upper()+'_'+str(sku)+'_SKUs_TID_'+track_id+'.xlsm'
    # try:
        # wb.win32com.client.save()
    # except:
    wb.Save()
    wb.Close()

        # Quit Excel
    xl.Quit()
    # Close the Excel application
    xl.Application.Quit()

    return 'CDS_LINESHEET_'+str(brand).upper()+'_'+str(sku)+'_SKUs_TID_'+track_id+'.xlsm'

if __name__ == '__main__':
    import sys
    function_name = sys.argv[1]
    args = sys.argv[2:]
    try:
        func = globals()[function_name]
        print(func(*args))

    except KeyError:
        print('Unknown function name:', function_name)