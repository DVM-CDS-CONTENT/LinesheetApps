#--- Generate
# #### Function Hear
# import mysql.connector
from sqlalchemy import create_engine
import pandas as pd
import openpyxl

# ===========
def get_string_location(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return '$'+openpyxl.utils.get_column_letter(column) + '$'+str(row)
# Add the columns to the worksheet, starting at cell A1
# ===========
def add_column(row,color,background,lists,wrapText,border,comment,author,allow_comment):
    for i, value in enumerate(lists):
        # Identity cell
        cell = worksheet.cell(row=row, column=i+1)
        cell.value = value
        # Setup cell font color
        cell.font = openpyxl.styles.Font(color=color)
        # Setup cell background color
        cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=background)
        # Setup Alignment at center for horizontal and vertical
        cell.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=wrapText)
        # Add a comment
        if comment[i] is not None and allow_comment==True:
            cell.comment = openpyxl.comments.Comment(text=comment[i], author='script')
        # Setup border
        if border == True or border == "True":
            cell.border = openpyxl.styles.Border(
                top=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                right=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                bottom=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                left=openpyxl.styles.Side(color='F0F0F0', style='thin')
            )
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].auto_size
# Add the columns to the worksheet, starting at cell A1
# ===========
def add_header(row,color,background,lists,bold,prefix = '',merge_cell = False,border=True):
    wrapText = False
    for i, value in enumerate(lists):
        # Identity cell
        cell = worksheet.cell(row=row, column=i+1)
        # Setup cell value
        if lists[i] != lists[i-1] :
            cell.value = str(prefix) + str(value)
            #merge
            if merge_group[i] == True and merge_cell == True:
                # check cell in same group
                num_cell = lists.count(value)
                # merge
                target_string = get_string_location(row-1, i) + ':'+ get_string_location(row, i+num_cell-1)
                workbook['IM_FORM'].merge_cells(target_string)
                wrapText = True
            else:
                wrapText = False
             # Setup border
            if border == True:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(color='F0F0F0', style='thin')
                )
                worksheet.cell(row=row+1, column=i+1).border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(color='F0F0F0', style='thin')
                )
        # Setup cell font color
        cell.font = openpyxl.styles.Font(color=color,bold=bold)
        # Setup cell background color
        cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=background)
        # Setup Alignment at center for horizontal and vertical
        cell.alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center',wrapText=wrapText,indent=1)
# ===========
def add_background_color(workbook, sheet_name, start_column, end_column, row, color):
    # Select the sheet where you want to add the background color
    sheet = workbook[sheet_name]
    # Set the color of the cells
    for col in range(start_column, end_column + 1):
        cell = sheet.cell(column=col, row=row)
        cell.fill = openpyxl.styles.PatternFill(fgColor=color, patternType='solid')
# ===========
def create_define_name(workbook ,sheet_name, attribute_code,length_list,column_num):
    string_range = get_string_location(1,column_num)+':'+get_string_location(length_list,column_num)
    new_range = openpyxl.workbook.defined_name.DefinedName(attribute_code, attr_text=sheet_name+'!'+string_range)

    try:
        workbook.defined_names.append(new_range)
    except:
        workbook.defined_names.add(new_range)


# ===========
def add_list_of_attribute(workbook, sheet_name,option_lists,attribute_code):
    sheet = workbook[sheet_name]
    col = sheet.max_column + 1
    sheet.cell(column=col, row=1, value=attribute_code)
    for i ,value in enumerate(option_lists):
        sheet.cell(column=col, row=i+2, value=value)
    #create definition name
    create_define_name(workbook ,sheet_name, attribute_code,len(option_lists),col-1)
# ===========
def add_dropdown(workbook, sheet_name, column, start_row, end_row, options):
    # Select the sheet where you want to add the dropdown
    sheet = workbook[sheet_name]
    # Create a DataValidation object with the dropdown options
    validation = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1='='+options, allow_blank=True)
    # Add the DataValidation object to the sheet
    sheet.add_data_validation(validation)
    # Add the DataValidation object to the specified range of cells
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(column=column, row=row)
        validation.add(cell)
# ===========
def add_formula(workbook, sheet_name, column, start_row, end_row, formula):
    # Select the sheet where you want to add the dropdown
    sheet = workbook[sheet_name]
    # Add the DataValidation object to the specified range of cells
    for row in range(start_row, end_row + 1):
        sheet.cell(column=column, row=row, value="="+str(formula))
        # add background
        cell = sheet.cell(column=column, row=row)
        cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='E7E7E7')
        cell.border = openpyxl.styles.Border(
                        top=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                        right=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                        bottom=openpyxl.styles.Side(color='F0F0F0', style='thin'),
                        left=openpyxl.styles.Side(color='F0F0F0', style='thin')
                    )
        # Setup Alignment at center for horizontal and vertical
        cell.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=False)
import re
import string
import datetime
import random
# ===========
def generate_track_id():
    current_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    random_number = random.randint(0, 999)
    track_id = current_time + str(random_number).zfill(3)
    return track_id
# ===========
def clean_string(string):
    """Replace special characters and convert a string to lowercase."""
    string = string.lower()
    return string
# ===========
def convert_sale_online_query(sale_channel):
    if sale_channel=='offline':
            return  "and sale_channel like '%offline%'"
    elif sale_channel=='online':
            return  "and sale_channel like '%online%'"
    else :
            return  ""

#--- Generate functions
# ===========
def generate_form(brand,template,sku,launch_date,stock_source,sale_channel,production_type):

    track_id = generate_track_id()
    # ### Read data ####################################################
    # this session use from get a configurable data from the database , so this will include
    # 1. attribute_setting
    # 2. attribute option setting
    # replace to multi template  #####################################

    filter_template_cat_sql = template.replace(",", "','")
    filter_template_header_sql = template.replace(",", " not in ('N','AR','AO') or ")
    selected_template_list = list(map(str,template.split(",")))

    # - Connect to the database
    # cnx = mysql.connector.connect(user='data_studio', password='a417528639', host='156.67.217.3', database='im_form')
    engine = create_engine('mysql+mysqlconnector://data_studio:a417528639@156.67.217.3/im_form')

    # - get header linesheet
    query = "SELECT * FROM im_form.attribute_setting where status = 'Actived' "+convert_sale_online_query(sale_channel)+" and ("+filter_template_header_sql+" not in ('N','AR','AO'))"
    attribute = pd.read_sql_query(query, engine)
    # attribute = pd.read_sql(query, cnx)
    attribute = attribute.drop_duplicates(subset=['linesheet_code'])
    # - get dropdown
    query = 'SELECT DISTINCT linesheet_code, input_option , option_en FROM u749625779_cdscontent.pim_attr_convert_option_lu'
    attribute_options = pd.read_sql_query(query, engine)
    lookup_sheet = attribute_options[['linesheet_code','input_option','option_en']]
    attribute_options = attribute_options[['linesheet_code','input_option']]
    attribute_options = attribute_options.drop_duplicates()
    # - get categories
    query = "SELECT label_th FROM im_form.categories_setting where deepen_level = 1 and family in ('"+filter_template_cat_sql +"')"

    categories_setting = pd.read_sql_query(query, engine)
    # print(categories_setting)
    # - Merge categories to options
    categories_setting['linesheet_code'] = 'online_categories'
    # - Rename the column headers
    categories_setting = categories_setting.rename(columns={'label_th': 'input_option'})
    attribute_options = pd.concat([attribute_options, categories_setting], axis=0)

    # - Close the connection
    # cnx.close()
# #### @ Create an config backup file ##########################3
    import os
    # # Create a folder for config
    # os.makedirs("config" , exist_ok=True)
    os.makedirs("linesheet" , exist_ok=True)
    # # backup config to local
    # attribute.to_json('config/attribute_config.json' , orient="records")
    # attribute_options.to_json('config/attribute_option_config.json' , orient="records")
    global codes
    global labels
    global indicators
    global comment
    global group_header
    global group_sub_header
    global merge_group
    global translation
    global fill_type
#create column indicator
    attribute["column_indication"] = attribute.loc[:, selected_template_list].apply(lambda x: ' '.join(x.astype(str)), axis=1)
# convert Data Frame to a list
    codes = attribute['linesheet_code'].values.tolist()
    labels = attribute['field_label'].values.tolist()
    indicators = attribute['column_indication'].values.tolist()
    comment = attribute['tool_tips'].values.tolist()
    group_header = attribute['session'].values.tolist()
    group_sub_header = attribute['sub_session'].values.tolist()
    merge_group =  attribute['merge_group'].values.tolist()
    translation = attribute['both_language'].values.tolist()
    formula = attribute['formula'].values.tolist()
    fill_type = attribute['field_type'].values.tolist()
# ##### Add translations column
    i = 0
    for j, value in enumerate(translation):
        if value == 1 or value == True or value == "True":
            codes.insert(i,codes[i]+'_en')
            labels.insert(i,labels[i]+' [English]')
            codes[i+1] = codes[i+1]+'_th'
            labels[i+1] = labels[i+1]+' [Thai]'
            indicators.insert(i,indicators[i])
            comment.insert(i,comment[i])
            group_header.insert(i,group_header[i])
            group_sub_header.insert(i,group_sub_header[i])
            merge_group.insert(i,merge_group[i])
            formula.insert(i,formula[i])
            fill_type.insert(i,fill_type[i])
            i += 1
        i += 1
# ### Create a sheet
    # This function use for create a excel file , so this one include the sheet below
    global workbook
    global worksheet
    global worksheet_form
    global ws_att_option
    global ws_inlink_data
    # - Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    workbook.properties.author = "Copyright © Central Online"
    # Change the name of the worksheet
    worksheet.title = 'IM_FORM'
    worksheet_form = workbook['IM_FORM']
    # Create a new sheet with the name "Sheet2"
    ws_att_option = workbook.create_sheet("ATT_OPTION")
    ws_inlink_data = workbook.create_sheet("IN_LINK_DATA")
    ws_lookup_sheet = workbook.create_sheet("LOOKUP_SHEET")
# #### General information's
# - Add version information
    worksheet.cell(row=2, column=2).value = 'New Omni linesheet Version ONE'
    worksheet.cell(row=3, column=2).value = 'Template : ' +str(template)
    worksheet.cell(row=2, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=True)
    worksheet.cell(row=3, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)
# Add lookup sheet
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(lookup_sheet, index=True, header=True):
        ws_lookup_sheet.append(r)
# Add inLink Data
    # -- get current date time
    from datetime import datetime
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    ws_inlink_data.append(['attribute', 'value'])
    ws_inlink_data.append(['brand', brand])
    ws_inlink_data.append(['sku', sku])
    ws_inlink_data.append(['stock_source', stock_source])
    ws_inlink_data.append(['template', template])
    ws_inlink_data.append(['sale_channel', sale_channel])
    ws_inlink_data.append(['launch_date', launch_date])
    ws_inlink_data.append(['production_type', production_type])
    ws_inlink_data.append(['form_version', '1.0.2'])
    ws_inlink_data.append(['create_date', dt_string])
    ws_inlink_data.append(['author', 'system'])
    ws_inlink_data.append(['track_id',track_id])
    ws_inlink_data.append(['check_require', 'unavailable'])
    ws_inlink_data.append(['check_list_validations', 'unavailable'])
# ##### Add a Table information's
    # Set the table range
    range_string = 'B4:C7'
    # Create a new table
    table = openpyxl.worksheet.table.Table(displayName="general_information", ref=range_string)
    # Add the table to the worksheet
    worksheet_form.add_table(table)
    # add table information
    ## add column
    worksheet_form.cell(row=4, column=2).value = 'General information'
    worksheet_form.cell(row=4, column=3).value = 'Value'
    ## add Row launch date
    worksheet_form.cell(row=5, column=2).value = 'Launch date'
    worksheet_form.cell(row=5, column=3).value = '=IN_LINK_DATA!B7' #launch_date
    ## add Row Production type
    worksheet_form.cell(row=6, column=2).value = 'Production type'
    worksheet_form.cell(row=6, column=3).value = '=IN_LINK_DATA!B8' #production_type
    ## add stock_source stock
    worksheet_form.cell(row=7, column=2).value = 'stock_source'
    worksheet_form.cell(row=7, column=3).value = '=IN_LINK_DATA!B4' #stock_source
    worksheet_form.cell(row=4, column=2).font = openpyxl.styles.Font(color='B41010',bold=False)
    worksheet_form.cell(row=4, column=3).font = openpyxl.styles.Font(color='B41010',bold=False)
    worksheet_form.cell(row=5, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)
    worksheet_form.cell(row=5, column=3).font = openpyxl.styles.Font(color='FFFF25',bold=False)
    worksheet_form.cell(row=6, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)
    worksheet_form.cell(row=6, column=3).font = openpyxl.styles.Font(color='FFFF25',bold=False)
    worksheet_form.cell(row=7, column=2).font = openpyxl.styles.Font(color='F0F0F0',bold=False)
    worksheet_form.cell(row=7, column=3).font = openpyxl.styles.Font(color='FFFF25',bold=False)
# #### Convert require field to indicator
    # change indicator
    for i in range(len(indicators)):
        if "R" in indicators[i]:
            indicators[i] ="Require ✔"
        else:
            indicators[i] = ""
        # indicators[i] = indicators[i].replace("R", "✔")
        # indicators[i] = indicators[i].replace("R", "Require")
        # indicators[i] = indicators[i].replace("O", "")
    # change boolean
    for i in range(len(merge_group)):
        if merge_group[i] == 1 or merge_group[i]  == True or merge_group[i] =="True":
            merge_group[i] = True
        else:
            merge_group[i] = False
# #### call function Add a column
    # call add column function
    add_column(1,'C91818','B41010',codes,False,False,comment,'script',False)
    add_column(8,'C91818','B41010',fill_type,False,False,comment,'script',False)
    add_column(12,'434343','E5E5E5',labels,True,True,comment,'script',True)
    add_column(11,'F73434','BFCAD0',indicators,False,True,comment,'script',False)
    # Get the length of the list
    latest_column = len(codes)
    # #### Add header
    add_header(9,'F0F0F0','161616',group_header,True,merge_cell=True,border=True)
    add_header(10,'F0F0F0','161616',group_sub_header,False,'▶ ',merge_cell=True,border=False)
    # #### Add row background space
    start_row = 1
    end_row = 8
    # call function
    for row in range(start_row,end_row+1):
        add_background_color(workbook, 'IM_FORM', 1, latest_column, row, 'B41010')
    # #### Set the width of columns
    # Set the width of all columns to fit the contents of the cells
    for column_index in range(1,  latest_column+ 1):
        column = openpyxl.utils.get_column_letter(column_index)
        worksheet.column_dimensions[column].auto_size = True
# # Add a dropdown list
    # - Loop add a dropdown
    # - Loop check value in each column and create a list of option
    start_col = 1
    for col in range(start_col, latest_column + 1):
        column_code = worksheet.cell(column=col, row=1).value
        option_frame = attribute_options[attribute_options['linesheet_code']==column_code]
        option_lists = option_frame['input_option'].values.tolist()
        if len(option_lists) != 0 :
            # Update the list of attribute in the sheet
            add_list_of_attribute(workbook, "ATT_OPTION", option_lists,column_code)
            # add data validation and create a dropdown
            add_dropdown(workbook,'IM_FORM', col, 13, int(sku)+13, column_code)
            option_lists.clear()
    #add formula
    for col in range(start_col, latest_column + 1):
        if formula[col-1]!=""  and formula[col-1]!=None:
           add_formula(workbook,'IM_FORM', col, 13, int(sku)+13, formula[col-1])
# ##### Save a file
    file_name = 'linesheet/'+'CDS_LINESHEET_'+str(brand).upper()+'_'+str(sku)+'_SKUs_TID_'+track_id+'.xlsm'
    workbook.save(file_name)
    from openpyxl import Workbook
    from openpyxl import load_workbook
    wb1 = load_workbook(file_name)
    wb2 = load_workbook(file_name, keep_vba=True)
    wb2.save(file_name)
# Add vba multi-select
    import os
    current_directory = os.getcwd()
    # current_directory ='linesheet'
    files = os.listdir(current_directory)
    import win32com.client
    # Open the Excel file
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.AutomationSecurity =1
    # xl.AutomationSecurity = 3

    wb = xl.Workbooks.Open(current_directory+'\\linesheet\\CDS_LINESHEET_'+str(brand).upper()+'_'+str(sku)+'_SKUs_TID_'+track_id+'.xlsm')
    # Access the sheet
    sheet = wb.Sheets("IM_FORM")
    # Access the VBA project
    vb = wb.VBProject
    # Create a new module
    module = vb.VBComponents.Add(1)
    module.Name = "im_form"
    code_name = sheet.codeName
    excelModule = wb.VBProject.VBComponents(code_name)
    # Insert the macro code
    import os



    with open(os.path.join(os.path.dirname(__file__), 'multi_select_cell.vba'), 'r') as f:
        code = f.read()
    # module.CodeModule.AddFromString(code)
    excelModule.CodeModule.AddFromString(code)
    # Add the macro to the sheet
    # sheet.CodeName = "Sheet1"
    sheet.OnSheetActivate = "im_form.Worksheet_Change"
    filename = 'linesheet/'+'CDS_LINESHEET_'+str(brand).upper()+'_'+str(sku)+'_SKUs_TID_'+track_id+'.xlsm'
    # try:
        # wb.win32com.client.save()
    # except:
    wb.Save()
    wb.Close()

        # Quit Excel
    xl.Quit()
    # Close the Excel application
    xl.Application.Quit()

    return 'CDS_LINESHEET_'+str(brand).upper()+'_'+str(sku)+'_SKUs_TID_'+track_id+'.xlsm'

if __name__ == '__main__':
    import sys
    function_name = sys.argv[1]
    args = sys.argv[2:]
    try:
        func = globals()[function_name]
        print(func(*args))

    except KeyError:
        print('Unknown function name:', function_name)